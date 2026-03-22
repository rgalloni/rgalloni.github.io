import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill
from glob import glob

def crea_excel_sintesi():
    """
    Unisce tutti i file CSV di sintesi in un unico file Excel,
    formattando le righe con colori alternati per ogni giorno.
    """
    sintesi_dir = "Relazione_4808/Itinerario_Giornaliero/sintesi/"
    output_excel_path = "sintesi_itinerario.xlsx"

    # Trova tutti i file CSV e li ordina
    csv_files = sorted(glob(os.path.join(sintesi_dir, "Giorno_*.csv")))

    if not csv_files:
        print(f"Nessun file CSV trovato nella directory: {sintesi_dir}")
        return

    # Legge e concatena tutti i file CSV
    df_list = [pd.read_csv(file, sep=';') for file in csv_files]
    full_df = pd.concat(df_list, ignore_index=True)

    # Crea un nuovo workbook e worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Sintesi Itinerario"

    # Scrive il DataFrame nel worksheet
    for r in dataframe_to_rows(full_df, index=False, header=True):
        ws.append(r)

    # Definisce i colori per le righe alternate
    color1 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid") # Bianco
    color2 = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid") # Grigio chiaro

    # Applica la formattazione
    current_day = None
    use_color1 = True
    
    # Salta l'header
    for row_index, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row)):
        day_value = row[0].value
        
        # Controlla se il giorno è cambiato
        if day_value and pd.notna(day_value) and day_value != current_day:
            current_day = day_value
            use_color1 = not use_color1
        
        # Applica il colore alla riga
        current_color = color1 if use_color1 else color2
        for cell in row:
            cell.fill = current_color

    # Adatta la larghezza delle colonne
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Salva il file Excel
    try:
        wb.save(output_excel_path)
        print(f"File Excel '{output_excel_path}' creato con successo.")
    except Exception as e:
        print(f"Errore durante il salvataggio del file Excel: {e}")

if __name__ == "__main__":
    crea_excel_sintesi()
